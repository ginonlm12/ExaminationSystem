import json
from django.shortcuts import render,redirect,reverse
from django.http import HttpResponse, JsonResponse
from . import forms,models
from openpyxl import Workbook
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from datetime import date, datetime, timedelta
from exam import models as QMODEL
from student import models as SMODEL
from exam import forms as QFORM
from django.contrib import messages
from exam.forms import QuestionForm  # Chỉnh sửa import
from exam.models import Question
from django.shortcuts import render, get_object_or_404, redirect
from exam.models import Course
from django.http import HttpResponse
from openpyxl import Workbook
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.contrib.auth.decorators import user_passes_test
from django.contrib import messages
from io import BytesIO

def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def edit_question_view(request, pk):
    # Lấy câu hỏi theo ID
    question = get_object_or_404(Question, id=pk)

    # Khởi tạo form với dữ liệu của câu hỏi
    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)

        if form.is_valid():
            # Lưu lại câu hỏi đã sửa
            question = form.save(commit=False)
            
            # Lưu các đáp án đúng dưới dạng JSON (danh sách các lựa chọn đúng)
            correct_answers = request.POST.getlist('correct_answers')
            question.correct_answers = correct_answers  # Lưu danh sách đáp án đúng dưới dạng JSON

            # Lưu câu hỏi vào cơ sở dữ liệu
            question.save()
            messages.success(request, "Câu hỏi đã được sửa thành công!")
            return redirect('teacher-view-question')  # Quay lại trang danh sách câu hỏi

    else:
        form = QuestionForm(instance=question)

    return render(request, 'teacher/edit_question.html', {'form': form})

#for showing signup/login button for teacher
def teacherclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return render(request,'teacher/teacherclick.html')

def teacher_signup_view(request):
    userForm=forms.TeacherUserForm()
    teacherForm=forms.TeacherForm()
    mydict={'userForm':userForm,'teacherForm':teacherForm}
    if request.method=='POST':
        userForm=forms.TeacherUserForm(request.POST)
        teacherForm=forms.TeacherForm(request.POST,request.FILES)
        if userForm.is_valid() and teacherForm.is_valid():
            user=userForm.save()
            user.set_password(user.password)
            user.save()
            teacher=teacherForm.save(commit=False)
            teacher.user=user
            teacher.save()
            my_teacher_group = Group.objects.get_or_create(name='TEACHER')
            my_teacher_group[0].user_set.add(user)
        return HttpResponseRedirect('teacherlogin')
    return render(request,'teacher/teachersignup.html',context=mydict)

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_dashboard_view(request):
    dict={
    
    'total_course':QMODEL.Course.objects.all().count(),
    'total_question':QMODEL.Question.objects.all().count(),
    'total_student':SMODEL.Student.objects.all().count()
    }
    return render(request,'teacher/teacher_dashboard.html',context=dict)

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_exam_view(request):
    return render(request,'teacher/teacher_exam.html')


@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_add_exam_view(request):
    courseForm = QFORM.CourseForm()

    if request.method == 'POST':
        courseForm = QFORM.CourseForm(request.POST)
        
        if courseForm.is_valid():
            # Lưu khóa thi vào cơ sở dữ liệu
            course = courseForm.save()

            # Sau khi tạo khóa thi, chuyển hướng đến trang thêm câu hỏi và truyền khóa thi vừa tạo vào
            return redirect('teacher-add-question', course_id=course.id)
        
        else:
            print("Form is invalid")

    return render(request, 'teacher/teacher_add_exam.html', {'courseForm': courseForm})


@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_exam_view(request):
    search = request.GET.get('search', '')  # Lấy từ input tìm kiếm trong form

    if search:
        # Tìm khóa thi theo tên khóa thi chứa từ khóa tìm kiếm
        courses = QMODEL.Course.objects.filter(course_name__icontains=search)
    else:
        courses = QMODEL.Course.objects.all()  # Nếu không có tìm kiếm, lấy tất cả khóa thi

    return render(request, 'teacher/teacher_view_exam.html', {'courses': courses})


@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def delete_exam_view(request,pk):
    course=QMODEL.Course.objects.get(id=pk)
    course.delete()
    return HttpResponseRedirect('/teacher/teacher-view-exam')

@login_required(login_url='adminlogin')
def teacher_question_view(request):
    return render(request,'teacher/teacher_question.html')

from django.contrib import messages
import openai  

@login_required
@user_passes_test(is_teacher)
def teacher_add_question_view(request, course_id=None):
    courses = Course.objects.all()
    course_name = None
    course = None
    
    if course_id:
        course = Course.objects.get(id=course_id)
        course_name = course.course_name
        
    questionForm = QuestionForm()

    if request.method == 'POST':
        questionForm = QuestionForm(request.POST)
        
        if questionForm.is_valid():
            # Lấy course_id từ form
            course_id_from_form = request.POST.get('courseID')  # Đảm bảo bạn lấy đúng 'courseID'
            if course_id_from_form:
                # Lấy đối tượng course từ course_id
                course = Course.objects.get(id=course_id_from_form)

            question = questionForm.save(commit=False)
            question.course = course  # Gán khóa thi bằng course_id
            question.save()
            messages.success(request, "Câu hỏi đã được lưu thành công!")
            return redirect('teacher-add-question', course_id=course.id if course else None)
        else:
            print(questionForm.errors)

    return render(request, 'teacher/teacher_add_question.html', {
        'questionForm': questionForm,
        'courses': courses,
        'course_id': course_id,
        'course_name': course_name,
        'course': course
    })
    
    
def recommend_answer(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        question = data.get('question')
        options = data.get('options')

        print("Received Question:", question)  # In câu hỏi nhận được
        print("Received Options:", options)  # In các lựa chọn nhận được

        # Tạo prompt cho mô hình LLM
        prompt = f"Question: {question}\nOptions: {options}\nWhich is the correct answer?"

        try:
            # Gọi API OpenAI (hoặc mô hình LLM khác)
            response = openai.Completion.create(
                engine="text-davinci-003",  # Bạn có thể sử dụng bất kỳ mô hình nào
                prompt=prompt,
                max_tokens=50,
                n=1,
                temperature=0.0,
            )

            answer = response.choices[0].text.strip()  # Lấy câu trả lời

            print("Recommended Answer:", answer)  # In đáp án gợi ý

            return JsonResponse({'correct_answer': answer})

        except Exception as e:
            print("Error during LLM request:", str(e))  # In lỗi nếu có
            return JsonResponse({'error': str(e)}, status=500)
        
@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def teacher_view_question_view(request):
    search = request.GET.get('search', '')  # Lấy giá trị từ form tìm kiếm

    if search:
        courses = Course.objects.filter(course_name__icontains=search)  # Lọc theo tên khóa thi
    else:
        courses = Course.objects.all()  # Nếu không tìm kiếm, lấy tất cả khóa thi

    return render(request, 'teacher/teacher_view_question.html', {'courses': courses})

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def see_question_view(request, pk):
    search = request.GET.get('search', '')  # Lấy giá trị tìm kiếm từ query params
    sort_order = request.GET.get('sort', 'asc')  # Lấy giá trị sắp xếp theo thứ tự ABC

    if search:
        questions = Question.objects.filter(course_id=pk, question__icontains=search)  # Tìm kiếm theo câu hỏi
    else:
        questions = Question.objects.filter(course_id=pk)

    if sort_order == 'asc':
        questions = questions.order_by('question')  # Sắp xếp theo tên câu hỏi tăng dần
    else:
        questions = questions.order_by('-question')  # Sắp xếp theo tên câu hỏi giảm dần

    course = Course.objects.get(id=pk)
    return render(request, 'teacher/see_question.html', {'questions': questions, 'course': course})

def export_questions_excel(request, pk):
    # Lấy tất cả câu hỏi của khóa thi
    questions = Question.objects.filter(course_id=pk)

    # Tạo workbook và sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Questions"

    # Thêm header vào sheet
    ws.append(['Câu hỏi', 'Điểm'])

    # Thêm dữ liệu câu hỏi vào sheet
    for question in questions:
        ws.append([question.question, question.marks])

    # Tạo response với file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="questions_{pk}.xlsx"'
    wb.save(response)

    return response


def export_question_excel(request):
    # Lấy tất cả câu hỏi
    questions = Question.objects.all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Câu hỏi"

    # Header
    ws.append(['Câu hỏi', 'Điểm'])

    for question in questions:
        ws.append([question.question, question.marks])

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="questions.xlsx"'
    wb.save(response)

    return response

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def remove_question_view(request,pk):
    question=QMODEL.Question.objects.get(id=pk)
    question.delete()
    return HttpResponseRedirect('/teacher/teacher-view-question')

from exam.models import Result
from exam.models import Course

@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
@login_required
def view_student_scores(request, course_id):
    # Lấy khóa thi
    course = Course.objects.get(id=course_id)

    # Lấy kết quả của tất cả học sinh đã làm bài thi này và xếp hạng theo điểm
    results = Result.objects.filter(exam=course).select_related('student').order_by('-marks', 'date')

    # Tính điểm cao nhất trong lượt thi đầu tiên và điểm cao nhất trong các lượt thi
    first_attempt_marks = {}
    highest_marks = {}

    # Thêm chức năng tìm kiếm theo tên học sinh
    search_query = request.GET.get('search', '')
    if search_query:
        results = results.filter(student__user__first_name__icontains=search_query)

    for result in results:
        student_name = f"{result.student.user.first_name} {result.student.user.last_name}"
        
        # Xác định điểm cao nhất trong lượt thi đầu tiên
        # Ví dụ: giả sử 'date' là trường lưu ngày làm bài, bạn có thể sử dụng một cách tính khác để xác định lượt thi đầu tiên.
        if result.student.user.first_name not in first_attempt_marks:
            first_attempt_marks[student_name] = result.marks
        
        # Tính điểm cao nhất trong tất cả các lượt thi
        if student_name not in highest_marks or result.marks > highest_marks[student_name]:
            highest_marks[student_name] = result.marks

    return render(request, 'teacher/view_student_scores.html', {
        'course': course,
        'results': results,
        'first_attempt_marks': first_attempt_marks,
        'highest_marks': highest_marks,
        'search_query': search_query  # Truyền search_query vào template để giữ giá trị trong ô tìm kiếm
    })

def export_to_excel(results):
    # Tạo workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Scores"

    # Đặt tiêu đề cột
    ws.append(['Tên học viên', 'Điểm', 'Ngày làm bài'])

    # Thêm dữ liệu vào Excel
    for result in results:
        ws.append([f"{result.student.user.first_name} {result.student.user.last_name}", result.marks, result.date])

    # Tạo phản hồi HTTP với file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="student_scores.xlsx"'

    # Lưu workbook vào phản hồi HTTP
    wb.save(response)
    return response

@login_required
@user_passes_test(is_teacher)
def student_marks_view(request, course_id):
    course = QMODEL.Course.objects.get(id=course_id)
    
    # Lấy tất cả kết quả cho khóa thi
    results = Result.objects.filter(exam=course)
    
    # 1. Lấy điểm trong lượt đầu tiên của mỗi học sinh
    first_attempt_results = results.order_by('student', 'date')
    first_attempt_marks = {}

    for result in first_attempt_results:
        # Chỉ lấy điểm đầu tiên của mỗi học sinh
        if result.student.id not in first_attempt_marks:
            first_attempt_marks[result.student.id] = result.marks

    # 2. Lấy điểm cao nhất trong các lượt thi của mỗi học sinh
    highest_marks = results.values('student').annotate(max_marks=Max('marks'))
    
    # Lấy danh sách học sinh và điểm cao nhất
    highest_marks_list = [(result['student'], result['max_marks']) for result in highest_marks]

    # Gửi dữ liệu vào template
    return render(request, 'teacher/student_marks_view.html', {
        'results': results,
        'course': course,
        'first_attempt_marks': first_attempt_marks,
        'highest_marks': highest_marks_list
    })
    
@login_required(login_url='teacherlogin')
@user_passes_test(is_teacher)
def question_statistics(request, course_id):
    # Lấy khóa thi
    course = Course.objects.get(id=course_id)

    # Lấy tất cả câu hỏi của bài thi
    questions = QMODEL.Question.objects.filter(course=course)

    question_stats = []

    for q in questions:
        # Lấy các kết quả câu hỏi từ bảng QuestionResult
        correct_count = QMODEL.QuestionResult.objects.filter(question=q, is_correct=True).count()
        incorrect_count = QMODEL.QuestionResult.objects.filter(question=q, is_correct=False).count()

        question_stats.append({
            'question': q.question,
            'correct_count': correct_count,
            'incorrect_count': incorrect_count,
            'total_count': correct_count + incorrect_count
        })

    return render(request, 'teacher/question_statistics.html', {
        'course': course,
        'question_stats': question_stats
    })
    
def export_excel(request):
    courses = Course.objects.all()

    # Tạo một workbook mới và thêm sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Khóa Thi"

    # Thêm header vào sheet
    ws.append(['Khóa thi', 'Tổng số câu hỏi', 'Tổng điểm'])

    # Thêm dữ liệu khóa thi vào sheet
    for course in courses:
        ws.append([course.course_name, course.question_number, course.total_marks])

    # Tạo response với file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="courses.xlsx"'
    wb.save(response)

    return response