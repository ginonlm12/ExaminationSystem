from django.shortcuts import render,redirect,reverse
from . import forms,models
from django.db.models import Sum
from django.contrib.auth.models import Group
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required,user_passes_test
from django.conf import settings
from datetime import date, timedelta
from django.db.models import Q
from django.core.mail import send_mail
from teacher import models as TMODEL
from student import models as SMODEL
from teacher import forms as TFORM
from student import forms as SFORM
from django.contrib.auth.models import User



def home_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')  
    return render(request,'exam/index.html')


def is_teacher(user):
    return user.groups.filter(name='TEACHER').exists()

def is_student(user):
    return user.groups.filter(name='STUDENT').exists()

def afterlogin_view(request):
    if is_student(request.user):      
        return redirect('student/student-dashboard')
                
    elif is_teacher(request.user):
        accountapproval=TMODEL.Teacher.objects.all().filter(user_id=request.user.id,status=True)
        if accountapproval:
            return redirect('teacher/teacher-dashboard')
        else:
            return render(request,'teacher/teacher_wait_for_approval.html')
    else:
        return redirect('admin-dashboard')



def adminclick_view(request):
    if request.user.is_authenticated:
        return HttpResponseRedirect('afterlogin')
    return HttpResponseRedirect('adminlogin')


@login_required(login_url='adminlogin')
def admin_dashboard_view(request):
    dict={
    'total_student':SMODEL.Student.objects.all().count(),
    'total_teacher':TMODEL.Teacher.objects.all().filter(status=True).count(),
    'total_course':models.Course.objects.all().count(),
    'total_question':models.Question.objects.all().count(),
    }
    return render(request,'exam/admin_dashboard.html',context=dict)

@login_required(login_url='adminlogin')
def admin_teacher_view(request):
    dict={
    'total_teacher':TMODEL.Teacher.objects.all().filter(status=True).count(),
    'pending_teacher':TMODEL.Teacher.objects.all().filter(status=False).count(),
    'salary':TMODEL.Teacher.objects.all().filter(status=True).aggregate(Sum('salary'))['salary__sum'],
    }
    return render(request,'exam/admin_teacher.html',context=dict)

@login_required(login_url='adminlogin')
def admin_view_teacher_view(request):
    search_query = request.GET.get('search', '')
    teachers = TMODEL.Teacher.objects.all().filter(status=True)

    if search_query:
        teachers = teachers.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(mobile__icontains=search_query) |
            Q(address__icontains=search_query)
        )

    return render(request, 'exam/admin_view_teacher.html', {'teachers': teachers})

import openpyxl
from django.http import HttpResponse

def export_teacher_excel(request):
    # Lấy danh sách giảng viên
    teachers = TMODEL.Teacher.objects.filter(status=True)

    # Tạo workbook và sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Giảng viên"

    # Tạo tiêu đề cột
    columns = ['Tên giảng viên', 'Ảnh đại diện', 'Liên hệ', 'Địa chỉ']
    for col_num, column_title in enumerate(columns, 1):
        ws.cell(row=1, column=col_num, value=column_title)

    # Điền dữ liệu vào các dòng
    for row_num, teacher in enumerate(teachers, 2):
        ws.cell(row=row_num, column=1, value=teacher.get_name)
        ws.cell(row=row_num, column=2, value=teacher.profile_pic.url)  # URL ảnh đại diện
        ws.cell(row=row_num, column=3, value=teacher.mobile)
        ws.cell(row=row_num, column=4, value=teacher.address)

    # Tạo response và gắn header để tải file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=teachers_list.xlsx'

    # Lưu file vào response
    wb.save(response)
    return response

from django.contrib.auth.hashers import make_password
from django.contrib import messages
from django.shortcuts import render, redirect
from django.shortcuts import get_object_or_404, redirect

@login_required(login_url='adminlogin')
def update_teacher_view(request, pk):
    teacher = get_object_or_404(TMODEL.Teacher, id=pk)
    user = teacher.user  # Lấy user của giáo viên

    userForm = TFORM.TeacherUserForm(instance=user)
    teacherForm = TFORM.TeacherForm(instance=teacher)
    
    print(teacherForm)

    if request.method == 'POST':
        userForm = TFORM.TeacherUserForm(request.POST, instance=user)
        teacherForm = TFORM.TeacherForm(request.POST, request.FILES, instance=teacher)

        if userForm.is_valid() and teacherForm.is_valid():
            user = userForm.save()
            if userForm.cleaned_data['password']:
                user.set_password(userForm.cleaned_data['password'])  # Cập nhật lại mật khẩu
                user.save()
            teacherForm.save()

            messages.success(request, "Cập nhật thông tin giáo viên thành công!")
            return redirect('admin-view-teacher')

    return render(request, 'exam/update_teacher.html', {
        'userForm': userForm,
        'teacherForm': teacherForm,
        'teacher': teacher
    })

@login_required(login_url='adminlogin')
def update_teacher_password(request, pk):
    teacher = get_object_or_404(TMODEL.Teacher, id=pk)
    user = teacher.user  # Lấy user của giáo viên

    if request.method == 'POST':
        # Lấy dữ liệu từ form mật khẩu
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')

        # Kiểm tra mật khẩu và xác nhận mật khẩu có giống nhau không
        if password != password_confirm:
            messages.error(request, "Mật khẩu và xác nhận mật khẩu không khớp.")
            return render(request, 'exam/update_teacher_password.html', {'teacher': teacher})

        # Cập nhật mật khẩu mới
        user.set_password(password)
        user.save()

        # Cập nhật session authentication hash sau khi thay đổi mật khẩu
        # update_session_auth_hash(request, user)

        messages.success(request, "Mật khẩu đã được thay đổi thành công!")
        return redirect('admin-view-teacher')  # Chuyển hướng tới danh sách giáo viên

    return render(request, 'exam/update_teacher_password.html', {
        'teacher': teacher
    })

@login_required(login_url='adminlogin')
def delete_teacher_view(request,pk):
    teacher=TMODEL.Teacher.objects.get(id=pk)
    user=User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return HttpResponseRedirect('/admin-view-teacher')

@login_required(login_url='adminlogin')
def admin_view_pending_teacher_view(request):
    teachers= TMODEL.Teacher.objects.all().filter(status=False)
    return render(request,'exam/admin_view_pending_teacher.html',{'teachers':teachers})


@login_required(login_url='adminlogin')
def approve_teacher_view(request,pk):
    teacherSalary=forms.TeacherSalaryForm()
    if request.method=='POST':
        teacherSalary=forms.TeacherSalaryForm(request.POST)
        if teacherSalary.is_valid():
            teacher=TMODEL.Teacher.objects.get(id=pk)
            teacher.salary=teacherSalary.cleaned_data['salary']
            teacher.status=True
            teacher.save()
        else:
            print("form is invalid")
        return HttpResponseRedirect('/admin-view-pending-teacher')
    return render(request,'exam/salary_form.html',{'teacherSalary':teacherSalary})

@login_required(login_url='adminlogin')
def reject_teacher_view(request,pk):
    teacher=TMODEL.Teacher.objects.get(id=pk)
    user=User.objects.get(id=teacher.user_id)
    user.delete()
    teacher.delete()
    return HttpResponseRedirect('/admin-view-pending-teacher')

@login_required(login_url='adminlogin')
def admin_view_teacher_salary_view(request):
    teachers= TMODEL.Teacher.objects.all().filter(status=True)
    return render(request,'exam/admin_view_teacher_salary.html',{'teachers':teachers})




@login_required(login_url='adminlogin')
def admin_student_view(request):
    dict={
    'total_student':SMODEL.Student.objects.all().count(),
    }
    return render(request,'exam/admin_student.html',context=dict)

@login_required(login_url='adminlogin')
def admin_view_student_view(request):
    # Lấy danh sách học viên
    students = SMODEL.Student.objects.all()

    # Tìm kiếm học viên theo tên, liên hệ, địa chỉ
    search_query = request.GET.get('search')
    if search_query:
        students = students.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(mobile__icontains=search_query) |
            Q(address__icontains=search_query)
        )

    return render(request, 'exam/admin_view_student.html', {'students': students})

import xlwt
# Export Excel cho danh sách học viên
def export_student_excel(request):
    # Tạo response là một file Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="students_list.xls"'

    wb = xlwt.Workbook()
    ws = wb.add_sheet('Students')

    # Đặt tiêu đề cho các cột
    columns = ['Tên', 'Liên hệ', 'Địa chỉ', 'Ảnh đại diện']
    for col_num, column_title in enumerate(columns):
        ws.write(0, col_num, column_title)

    # Lấy dữ liệu học viên
    students = SMODEL.Student.objects.all()
    for row_num, student in enumerate(students, 1):
        ws.write(row_num, 0, f"{student.user.first_name} {student.user.last_name}")
        ws.write(row_num, 1, student.mobile)
        ws.write(row_num, 2, student.address)
        ws.write(row_num, 3, str(student.profile_pic))  # Hình ảnh

    wb.save(response)
    return response



@login_required(login_url='adminlogin')
def update_student_view(request,pk):
    student = get_object_or_404(SMODEL.Student, id=pk)
    user = student.user  # Lấy user của giáo viên
    
    userForm=SFORM.StudentUserForm(instance=user)
    studentForm=SFORM.StudentForm(instance=student)
    
    print(studentForm)
    
    if request.method == 'POST':
        userForm = SFORM.StudentUserForm(request.POST, instance=user)
        studentForm = SFORM.StudentForm(request.POST, request.FILES, instance=student)

        if userForm.is_valid() and studentForm.is_valid():
            user = userForm.save()
            if userForm.cleaned_data['password']:
                user.set_password(userForm.cleaned_data['password'])  # Cập nhật lại mật khẩu
                user.save()
            studentForm.save()

            messages.success(request, "Cập nhật thông tin học viên thành công!")
            return redirect('admin-view-student')

    return render(request, 'exam/update_student.html', {
        'userForm': userForm,
        'teacherForm': studentForm,
        'teacher': student
    })

@login_required(login_url='adminlogin')
def update_student_password(request, pk):
    student = get_object_or_404(SMODEL.Student, id=pk)
    user = student.user  # Lấy user của giáo viên

    if request.method == 'POST':
        # Lấy dữ liệu từ form mật khẩu
        password = request.POST.get('password')
        password_confirm = request.POST.get('password_confirm')

        # Kiểm tra mật khẩu và xác nhận mật khẩu có giống nhau không
        if password != password_confirm:
            messages.error(request, "Mật khẩu và xác nhận mật khẩu không khớp.")
            return render(request, 'exam/update_student_password.html', {'student': student})

        # Cập nhật mật khẩu mới
        user.set_password(password)
        user.save()

        # Cập nhật session authentication hash sau khi thay đổi mật khẩu
        # update_session_auth_hash(request, user)

        messages.success(request, "Mật khẩu đã được thay đổi thành công!")
        return redirect('admin-view-student')  # Chuyển hướng tới danh sách giáo viên

    return render(request, 'exam/update_student_password.html', {
        'student': student
    })

@login_required(login_url='adminlogin')
def delete_student_view(request,pk):
    student=SMODEL.Student.objects.get(id=pk)
    user=User.objects.get(id=student.user_id)
    user.delete()
    student.delete()
    return HttpResponseRedirect('/admin-view-student')


@login_required(login_url='adminlogin')
def admin_course_view(request):
    return render(request,'exam/admin_course.html')

from exam import forms as QFORM
@login_required(login_url='adminlogin')
def admin_add_course_view(request):
    courseForm = QFORM.CourseForm()

    if request.method == 'POST':
        courseForm = QFORM.CourseForm(request.POST)
        
        if courseForm.is_valid():
            # Lưu khóa thi vào cơ sở dữ liệu
            course = courseForm.save()

            # Sau khi tạo khóa thi, chuyển hướng đến trang thêm câu hỏi và truyền khóa thi vừa tạo vào
            return redirect('admin-add-question', course_id=course.id)
        
        else:
            print("Form is invalid")

    return render(request, 'exam/admin_add_course.html', {'courseForm': courseForm})


@login_required(login_url='adminlogin')
def admin_view_course_view(request):
    courses = models.Course.objects.all()
    return render(request,'exam/admin_view_course.html',{'courses':courses})

@login_required(login_url='adminlogin')
def delete_course_view(request,pk):
    course=models.Course.objects.get(id=pk)
    course.delete()
    return HttpResponseRedirect('/admin-view-course')



@login_required(login_url='adminlogin')
def admin_question_view(request):
    return render(request,'exam/admin_question.html')

from exam.models import Course
from exam.forms import QuestionForm 

@login_required(login_url='adminlogin')
def admin_add_question_view(request, course_id=None):
    courses = Course.objects.all()
    course_name = None
    course = None
    
    if course_id:
        course = Course.objects.get(id=course_id)
        course_name = course.course_name
        
    questionForm = QuestionForm()

    if request.method == 'POST':
        questionForm = QuestionForm(request.POST)
        
        if questionForm.is_valid():
            # Lấy course_id từ form
            course_id_from_form = request.POST.get('courseID')  # Đảm bảo bạn lấy đúng 'courseID'
            if course_id_from_form:
                # Lấy đối tượng course từ course_id
                course = Course.objects.get(id=course_id_from_form)

            question = questionForm.save(commit=False)
            question.course = course  # Gán khóa thi bằng course_id
            question.save()
            messages.success(request, "Câu hỏi đã được lưu thành công!")
            return redirect('admin-add-question', course_id=course.id if course else None)
        else:
            print(questionForm.errors)

    return render(request, 'exam/admin_add_question.html', {
        'questionForm': questionForm,
        'courses': courses,
        'course_id': course_id,
        'course_name': course_name,
        'course': course
    })

@login_required(login_url='adminlogin')
def admin_view_question_view(request):
    courses= models.Course.objects.all()
    return render(request,'exam/admin_view_question.html',{'courses':courses})

@login_required(login_url='adminlogin')
def view_question_view(request,pk):
    questions=models.Question.objects.all().filter(course_id=pk)
    return render(request,'exam/view_question.html',{'questions':questions})

@login_required(login_url='adminlogin')
def delete_question_view(request,pk):
    question=models.Question.objects.get(id=pk)
    question.delete()
    return HttpResponseRedirect('/admin-view-question')

@login_required(login_url='adminlogin')
def admin_view_student_marks_view(request):
    students= SMODEL.Student.objects.all()
    return render(request,'exam/admin_view_student_marks.html',{'students':students})

@login_required(login_url='adminlogin')
def admin_view_marks_view(request,pk):
    courses = models.Course.objects.all()
    response =  render(request,'exam/admin_view_marks.html',{'courses':courses})
    response.set_cookie('student_id',str(pk))
    return response

@login_required(login_url='adminlogin')
def admin_check_marks_view(request,pk):
    course = models.Course.objects.get(id=pk)
    student_id = request.COOKIES.get('student_id')
    student= SMODEL.Student.objects.get(id=student_id)

    results= models.Result.objects.all().filter(exam=course).filter(student=student)
    return render(request,'exam/admin_check_marks.html',{'results':results})
    




def aboutus_view(request):
    return render(request,'exam/aboutus.html')

def contactus_view(request):
    sub = forms.ContactusForm()
    if request.method == 'POST':
        sub = forms.ContactusForm(request.POST)
        if sub.is_valid():
            email = sub.cleaned_data['Email']
            name=sub.cleaned_data['Name']
            message = sub.cleaned_data['Message']
            send_mail(str(name)+' || '+str(email),message,settings.EMAIL_HOST_USER, settings.EMAIL_RECEIVING_USER, fail_silently = False)
            return render(request, 'exam/contactussuccess.html')
    return render(request, 'exam/contactus.html', {'form':sub})


